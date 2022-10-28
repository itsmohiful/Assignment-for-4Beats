import re
import time

#openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
#webdriver
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

#input targeted keyword
keyword = str(input("Enter Searching Keyword: "))


def scrap_goole_suggestion(keyword):
    PATH = "./selenium-webdriver/chromedriver.exe"
    driver = webdriver.Chrome(PATH)

    driver.get("https://www.google.com/")


    try:
        #find search field
        search = WebDriverWait(driver, 4).until(
            EC.presence_of_element_located((By.NAME,"q"))
        )

        #send targetd keyword
        search.send_keys(keyword)
        time.sleep(2)

        #fetch google suggestions
        search_lists = driver.find_elements(By.TAG_NAME, "li")

        data_list = []

        
        for li in search_lists:
            li_value = str(li.text)
            if li_value:
                data_list.append(li_value)

        #sorting data
        data_list.sort(key=len)
        last_index = len(data_list)-1

        shortest_option = data_list[0]
        longest_option = data_list[last_index]

        #write data in exel file
        wb = load_workbook('./data.xlsx')
        sheet_name = ["Saturday", "Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
        names = wb.sheetnames
        for sheet in sheet_name:
            if sheet in names:
                continue
            else:
                wb.create_sheet(sheet)
                

        day = time.strftime("%A")
        ws = wb[f'{day}']
        

        #exel styling with openpyxl
        ft = Font(bold=True, size=13)
        ws['A1'] = ' '
        ws['B1'] = ' '
        ws['C1'] = ' '
        ws['D1'] = 'Longest Option'
        ws['E1'] = 'Shortest Option'
        ws["A1"].font = ws["B1"].font = ws["D1"].font = ws["E1"].font = ft
        
        #append data in existing exel file
        ws.append(['Keyword ', keyword," ", longest_option, shortest_option])
        wb.save('./data.xlsx')

        print("Script run successfully")
        
        driver.quit()

    except:
        driver.quit()


scrap_goole_suggestion(keyword)






